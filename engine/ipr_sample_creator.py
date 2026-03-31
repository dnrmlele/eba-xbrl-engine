"""
engine/ipr_sample_creator.py
Creates a fully formatted fillable Excel workbook for SEPA_IPR data entry.
Mirrors the style conventions of create_sample_workbook() (LCR).
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine.definitions.ipr_templates import (
    TEMPLATE_STRUCTURE,
    TEMPLATE_FILE_NAMES,
    INDICATOR_MAP,
    EURO_AREA_FILED,
    IPR_TEMPLATES,
)

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FONT  = Font(color="FFFFFF", bold=True, size=10)
ROW_FILL      = PatternFill("solid", fgColor="D6E4F0")
INPUT_FILL    = PatternFill("solid", fgColor="FFFDE7")
DISABLED_FILL = PatternFill("solid", fgColor="EEEEEE")
META_FONT     = Font(italic=True, color="888888", size=9)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

NON_EURO_TEMPLATES = {"S 01.02", "S 02.02"}


def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN


def _write_template_sheet(ws, tpl_code: str, rows: list, cols: list,
                          year: int, disabled: bool) -> None:
    indicator = INDICATOR_MAP[tpl_code]

    # ── Row 1: meta banner ───────────────────────────────────────────────────
    ws.cell(1, 1, f"SEPA_IPR  |  {tpl_code}  |  Reference year: {year}"
                  f"  |  Filing indicator: {indicator}"
                  f"  |  File: {TEMPLATE_FILE_NAMES[tpl_code]}")
    ws.cell(1, 1).font = META_FONT
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=max(len(cols) + 1, 4))

    if disabled:
        ws.cell(2, 1, "⚠️  N/A for Euro-area PSPs — filing indicator set to FALSE")
        ws.cell(2, 1).font = Font(color="CC0000", bold=True, size=10)
        ws.cell(2, 1).fill = DISABLED_FILL
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2,   end_column=max(len(cols) + 1, 4))
        ws.column_dimensions["A"].width = 60
        return

    # ── Column groupings (hierarchical headers) ───────────────────────────────
    # Define visual groupings for better layout (template-specific)
    col_groupings = _get_column_groupings(tpl_code, cols)

    # ── Row 2: Column group headers (if any) ──────────────────────────────────
    if col_groupings:
        group_row = 2
        col_idx = 2
        for group_label, group_cols in col_groupings:
            span = len(group_cols)
            if span > 1:
                ws.merge_cells(start_row=group_row, start_column=col_idx,
                               end_row=group_row, end_column=col_idx + span - 1)
            cell = ws.cell(group_row, col_idx, group_label)
            cell.fill = SUBHEAD_FILL
            cell.font = SUBHEAD_FONT
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            col_idx += span
        ws.row_dimensions[group_row].height = 24

    # ── Row 3: column headers ────────────────────────────────────────────────
    h_row = 3
    ws.cell(h_row, 1, "Row Code & Description").fill = HEADER_FILL
    ws.cell(h_row, 1).font = HEADER_FONT
    ws.column_dimensions["A"].width = 55

    for i, (col_code, col_label) in enumerate(cols, start=2):
        c = ws.cell(h_row, i, f"{col_code}\n{col_label}")
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[h_row].height = 40

    cell_map = IPR_TEMPLATES.get(tpl_code, {}).get("cell_map", {})

    # ── Data rows ────────────────────────────────────────────────────────────
    for d_row_idx, (row_code, row_label) in enumerate(rows, start=h_row + 1):
        # Determine whether this row is a visual section header or a fillable data row.
        row_label_text = row_label.strip()
        is_total = row_code == "0010"

        is_section = False
        if "Breakdown by" in row_label_text:
            if " — " in row_label_text:
                # Data row in a breakdown section, keep value part only.
                _, row_label_text = row_label_text.split(" — ", 1)
                row_label_text = row_label_text.strip()
            else:
                # Pure section header row.
                is_section = True

        if is_section:
            label_text = row_label_text
        else:
            label_text = f"{row_code}  {row_label_text}".rstrip()

        label_cell = ws.cell(d_row_idx, 1, label_text)

        if is_section:
            label_cell.fill = PatternFill("solid", fgColor="D3D3D3")
            label_cell.font = Font(bold=True, size=11, color="333333")
            ws.merge_cells(start_row=d_row_idx, start_column=1,
                           end_row=d_row_idx, end_column=len(cols) + 1)
            label_cell.alignment = Alignment(indent=1, vertical="center")
            ws.row_dimensions[d_row_idx].height = 18
        else:
            label_cell.fill = ROW_FILL if is_total else PatternFill("solid", fgColor="EAF3FB")
            label_cell.font = Font(bold=is_total, size=10)
            label_cell.alignment = Alignment(indent=1 if is_total else 3, vertical="center")

        # Data cells
        for c_idx, (col_code, _) in enumerate(cols, start=2):
            inp = ws.cell(d_row_idx, c_idx)
            if not is_section and (row_code, col_code) in cell_map:
                inp.fill = INPUT_FILL
                inp.alignment = Alignment(horizontal="right", vertical="center")
                inp.number_format = "#,##0"
            else:
                inp.fill = DISABLED_FILL

    _apply_border(ws, h_row, h_row + len(rows), 1, len(cols) + 1)
    ws.freeze_panes = ws.cell(h_row + 1, 2)


def _get_column_groupings(tpl_code: str, cols: list):
    """
    Returns hierarchical column groupings for better visual organization.
    Format: [(group_label, [(col_code, col_label), ...]), ...]
    """
    col_codes = [code for code, _ in cols]
    
    # S 01.01 & S 01.02: Credit transfers sent vs received
    if tpl_code in {"S 01.01", "S 01.02"}:
        sent_cols = [(c, l) for c, l in cols if c in {"0010", "0020", "0030", "0040"}]
        recv_cols = [(c, l) for c, l in cols if c in {"0050", "0060", "0070", "0080"}]
        if sent_cols and recv_cols:
            return [
                ("Credit transfers sent", sent_cols),
                ("Credit transfers received", recv_cols),
            ]
    
    # S 02.01 & S 02.02: Charges sent vs received
    elif tpl_code in {"S 02.01", "S 02.02"}:
        sent_cols = [(c, l) for c, l in cols if c in {"0010", "0020"}]
        recv_cols = [(c, l) for c, l in cols if c in {"0030", "0040"}]
        if sent_cols and recv_cols:
            return [
                ("Charges for credit transfers sent", sent_cols),
                ("Charges for credit transfers received", recv_cols),
            ]
    
    # S 03.00: Multiple charge types
    elif tpl_code == "S 03.00":
        if len(cols) >= 2:
            return [
                ("Payment accounts & charges", cols),
            ]
    
    return None


def create_ipr_sample(entity_lei: str, years: list, is_euro_area: bool = True) -> bytes:
    """
    Returns bytes of a .xlsx workbook with one sheet per (year × template).
    Yellow cells = data entry. Greyed sheets = N/A for euro-area PSPs.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── INDEX sheet ──────────────────────────────────────────────────────────
    idx = wb.create_sheet("INDEX")
    idx.sheet_properties.tabColor = "1F4E79"
    idx["A1"] = "xBRL-CSV Generator — SEPA_IPR Sample Template"
    idx["A1"].font = Font(bold=True, size=14, color="1F4E79")
    idx["A3"] = f"LEI:              {entity_lei}"
    idx["A4"] = f"Reference years:  {', '.join(str(y) for y in sorted(years))}"
    idx["A5"] = f"Euro-area PSP:    {is_euro_area}"
    idx["A6"] = f"First remittance: 9 April 2026"
    idx["A8"] = "HOW TO USE:"
    idx["A8"].font = Font(bold=True)
    idx["A9"]  = "1. Fill yellow cells on each template sheet"
    idx["A10"] = "2. Save the workbook"
    idx["A11"] = "3. Upload in xBRL-CSV Generator → CSSF tab → Generate IPR ZIPs"
    idx["A12"] = "4. One ZIP will be produced per reference year"
    idx["A13"] = "   Submit each ZIP separately to CSSF eDesk"
    idx["A15"] = "TEMPLATES INCLUDED:"
    idx["A15"].font = Font(bold=True)
    for r, (tpl, (_, cols)) in enumerate(TEMPLATE_STRUCTURE.items(), start=16):
        ind = INDICATOR_MAP[tpl]
        na = " ← N/A (Euro-area PSP)" if is_euro_area and tpl in NON_EURO_TEMPLATES else ""
        idx.cell(r, 1, f"  {tpl}   {ind}{na}")
    idx.column_dimensions["A"].width = 65

    # ── Template sheets ──────────────────────────────────────────────────────
    for year in sorted(years):
        for tpl_code, (rows, cols) in TEMPLATE_STRUCTURE.items():
            disabled = is_euro_area and tpl_code in NON_EURO_TEMPLATES
            sheet_name = f"{year}_{tpl_code.replace(' ', '_')}"
            ws = wb.create_sheet(sheet_name)
            ws.sheet_properties.tabColor = "AAAAAA" if disabled else "1F4E79"
            _write_template_sheet(ws, tpl_code, rows, cols, year, disabled)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
