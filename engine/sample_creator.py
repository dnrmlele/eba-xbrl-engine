
"""Creates sample EBA-style Excel templates for C72-C76 with multi-currency."""
import io, random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .definitions.lcr_templates import LCR_TEMPLATES

BLUE   = PatternFill("solid", fgColor="003399")
BLUE2  = PatternFill("solid", fgColor="4472C4")
GREEN  = PatternFill("solid", fgColor="E2EFDA")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
ORANGE = PatternFill("solid", fgColor="FCE4D6")
DGREY  = PatternFill("solid", fgColor="D9E1F2")
WF = Font(bold=True, color="FFFFFF", name="Arial", size=9)
BF = Font(bold=True, name="Arial", size=9)
LA = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
RA = Alignment(horizontal="right",  vertical="center")
TB = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"))

# Scaling factors per currency (relative to EUR notional amounts)
CCY_SCALE = {"EUR": 1.0, "USD": 1.08, "GBP": 0.86, "CHF": 0.97,
             "JPY": 163.0, "SEK": 11.5, "NOK": 11.8}

def create_sample_workbook(currencies: list = None) -> io.BytesIO:
    """
    Create sample EBA workbook with sheets per template × currency.
    Sheet naming: "C 72.00" (EUR), "C 72.00 USD", "C 72.00 GBP", …
    """
    if currencies is None:
        currencies = ["EUR", "USD", "GBP"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for currency in currencies:
        scale = CCY_SCALE.get(currency, 1.0)
        seed  = abs(hash(currency)) % 1000
        for key, tdef in LCR_TEMPLATES.items():
            sheet_title = (tdef["sheet_name"] if currency == "EUR"
                           else f"{tdef['sheet_name']} {currency}")
            ws = wb.create_sheet(title=sheet_title)
            _build_sheet(ws, tdef, currency, scale, seed)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _build_sheet(ws, tdef, currency, scale, seed):
    random.seed(seed)
    cols, rows = tdef["columns"], tdef["rows"]
    ncols = len(cols)

    # Row 1: title
    ws.merge_cells(f"A1:{get_column_letter(2+ncols)}1")
    c = ws["A1"]
    c.value = f"{tdef['template_code']} – {tdef['name']}  [{currency}]"
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.fill = BLUE; c.alignment = CA
    ws.row_dimensions[1].height = 25

    # Rows 2-4: metadata
    ws["A2"]="Entity Name:"; ws["B2"]="SAMPLE BANK SA"
    ws["D2"]="Reference Date:"; ws["E2"]="2025-12-31"
    ws["A3"]="LEI:"; ws["B3"]="7LTWFZYICNSX8D621K86"
    ws["A4"]="Currency:"; ws["B4"]=currency
    ws["D4"]="Unit:"; ws["E4"]="thousands"
    for cell in ["A2","D2","A3","A4","D4"]:
        ws[cell].font = BF
    ws.row_dimensions[5].height = 6

    # Row 6: col headers
    ws["A6"]="Row"; ws["A6"].font=WF; ws["A6"].fill=BLUE; ws["A6"].alignment=CA
    ws["B6"]="Description"; ws["B6"].font=WF; ws["B6"].fill=BLUE; ws["B6"].alignment=CA
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=44
    for i,(cid,cd) in enumerate(cols.items(), start=3):
        cl = get_column_letter(i)
        c = ws.cell(row=6, column=i)
        c.value = f"{cd['label']}\n({cid})"
        c.font=WF; c.fill=BLUE; c.alignment=CA
        ws.column_dimensions[cl].width=18
    ws.row_dimensions[6].height=34

    # Row 7: col ID row
    ws["A7"].fill=BLUE2; ws["B7"].fill=BLUE2
    for i,(cid,_) in enumerate(cols.items(), start=3):
        c=ws.cell(row=7,column=i)
        c.value=cid; c.font=Font(bold=True,color="FFFFFF",name="Arial",size=8)
        c.fill=BLUE2; c.alignment=CA
    ws.row_dimensions[7].height=14

    # Data rows
    for rid, rdef in rows.items():
        er    = rdef["excel_row"]
        label = rdef["label"].upper()
        is_total   = "TOTAL" in label or rid in ("r0310","r0200")
        is_section = any(x in label for x in ["LEVEL 1","LEVEL 2A","LEVEL 2B"])
        fill = (DGREY if is_total
                else GREEN  if "LEVEL 1"  in label
                else YELLOW if "LEVEL 2A" in label
                else ORANGE if "LEVEL 2B" in label
                else PatternFill())

        ic = ws.cell(row=er, column=1)
        ic.value=rid.upper(); ic.font=Font(bold=is_total or is_section,name="Arial",size=8)
        ic.alignment=CA; ic.fill=fill; ic.border=TB

        dc = ws.cell(row=er, column=2)
        dc.value=rdef["label"]
        dc.font=Font(bold=is_total or is_section,italic=is_section,name="Arial",size=9)
        dc.alignment=LA; dc.fill=fill; dc.border=TB

        for i,(cid,_) in enumerate(cols.items(), start=3):
            vc = ws.cell(row=er, column=i)
            if is_section:
                vc.value = None
            elif is_total:
                vc.value = (int(random.randint(5000,50000)*1000*scale)
                            if cid != "c0020" else None)
                if vc.value: vc.number_format="#,##0"; vc.font=BF
            else:
                if cid == "c0020":
                    vc.value = random.choice([0.0, 0.15, 0.25, 0.35])
                    vc.number_format = "0.00%"
                else:
                    vc.value = (int(random.randint(100,5000)*1000*scale)
                                if random.random() > 0.3 else None)
                    if vc.value: vc.number_format="#,##0"
            vc.alignment=RA; vc.fill=fill; vc.border=TB

    ws.freeze_panes = "C8"
