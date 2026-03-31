"""
engine/ipr_reader.py
Reads a filled IPR Excel (produced by ipr_sample_creator.py).
Returns data_by_year: {year: {tpl_code: {(row_code, col_code): float}}}
"""

import re
import openpyxl

from engine.definitions.ipr_templates import TEMPLATE_STRUCTURE

# Sheet name pattern: "2022_S_01_01"  "2025_S_04.00"
SHEET_RE = re.compile(r"^(\d{4})_(S_\d{2}[_\.]\d{2})$")


def _sheet_key(sheet_name: str):
    """Returns (year: int, tpl_code: str) or (None, None)."""
    # Accept variations: 2025_S_01_01, 2025_S_01.01, S_01_01, S 01.01
    sheet_name = sheet_name.strip()
    m = SHEET_RE.match(sheet_name)
    if m:
        year = int(m.group(1))
        raw = m.group(2)
    else:
        # Try without year prefix; allow also 'S_01_01', 'S 01.01'
        raw = sheet_name.replace(' ', '_')
        if not raw.startswith('S_'):
            return None, None
        year = None

    raw = raw.replace('.', '_')
    parts = raw.split("_")
    if len(parts) < 3:
        return None, None
    tpl = f"{parts[0]} {parts[1]}.{parts[2]}"
    if tpl not in TEMPLATE_STRUCTURE:
        return None, None
    return year, tpl


def _extract_row_code(cell_value):
    if cell_value is None:
        return None
    val = str(cell_value).strip()
    m = re.search(r"(\d{4})", val)
    return m.group(1) if m else None


def _to_float(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(' ', '')
    s = s.replace(',', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def read_ipr_excel(file_obj, years: list) -> dict:
    """
    Returns {year: {tpl_code: {(row_code, col_code): float}}}.
    Skips disabled (N/A) sheets.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    data_by_year = {y: {} for y in years}

    for sheet_name in wb.sheetnames:
        year, tpl_code = _sheet_key(sheet_name)
        if tpl_code is None:
            continue
        if year is not None and year not in years:
            continue

        ws = wb[sheet_name]
        rows_def, cols_def = TEMPLATE_STRUCTURE[tpl_code]
        row_codes = {r[0] for r in rows_def}
        col_codes = [c[0] for c in cols_def]

        tpl_data = {}
        for excel_row in ws.iter_rows(values_only=True):
            if not excel_row or excel_row[0] is None:
                continue
            rc = _extract_row_code(excel_row[0])
            if rc is None or rc not in row_codes:
                continue
            for c_idx, col_code in enumerate(col_codes, start=1):
                raw = excel_row[c_idx] if c_idx < len(excel_row) else None
                val = _to_float(raw)
                if val is None:
                    continue
                tpl_data[(rc, col_code)] = val

        if tpl_data:
            target_year = year if year is not None else years[0]
            if target_year in data_by_year:
                data_by_year[target_year][tpl_code] = tpl_data

    wb.close()
    return data_by_year
