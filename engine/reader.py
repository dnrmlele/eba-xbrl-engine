""" 
engine/reader.py
EBA Excel reader — maps filled workbook cells to XBRL-CSV data points.
Uses cell_map from lcr_templates.py to resolve concept IDs at read time.
"""
import re
import openpyxl
from engine.definitions.lcr_templates import LCR_TEMPLATES

# Master template registry
_ALL_TEMPLATES = {**LCR_TEMPLATES}

# LCR sheet name pattern: "C 72.00" or "C 72.00 USD"
_LCR_RE = re.compile(
    r"^(C 7[2-6]\.\d{2}(?:\.[ab])?)(?:\s+([A-Z]{3}))?$"
)

def _template_key_from_code(code: str, sub: str) -> str:
    """Convert 'C 72.00' + 'a' → 'C_72.00.a'"""
    return code.replace(" ", "_").replace(".", "_") + f".{sub}"


class EBAExcelReader:
    def read(self, fileobj, default_currency: str = "EUR") -> dict:
        wb = openpyxl.load_workbook(fileobj, data_only=True)
        results = {}

        for sheet_name in wb.sheetnames:
            m = _LCR_RE.match(sheet_name.strip())
            if not m:
                continue
            code     = m.group(1)    # e.g. "C 72.00"
            ccy      = m.group(2) or default_currency
            sub      = "w" if ccy != default_currency else "a"
            tmpl_key = _template_key_from_code(code, sub)

            tmpl = _ALL_TEMPLATES.get(tmpl_key)
            if tmpl is None:
                continue

            ws          = wb[sheet_name]
            data_points = self._extract(ws, tmpl)

            results[sheet_name] = {
                "module":                "LCR",
                "template_key":          tmpl_key,
                "template_code":         code,
                "sub_template":          sub,
                "name":                  tmpl["name"],
                "currency":              ccy,
                "xbrl_filing_indicator": tmpl["xbrl_filing_indicator"],
                "cell_map":              tmpl["cell_map"],
                "data_points":           data_points,
            }

        wb.close()
        return results

    def _extract(self, ws, tmpl: dict) -> dict:
        """Extract cell values using row/col positions from template definition."""
        data_points = {}
        for row_code, row_meta in tmpl["rows"].items():
            er = row_meta.get("excel_row")
            if er is None:
                continue
            for col_code, col_meta in tmpl["columns"].items():
                ec = col_meta.get("excel_col")
                if ec is None:
                    continue
                val = ws.cell(row=er, column=ec).value
                if val is not None:
                    data_points[f"{row_code}_{col_code}"] = val
        return data_points
