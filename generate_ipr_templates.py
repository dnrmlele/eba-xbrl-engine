#!/usr/bin/env python3
"""
generate_ipr_templates.py
Extracts SEPA_IPR concept IDs from the EBA RF 4.2 DPM Annotated Excel
and updates engine/definitions/ipr_templates.py with real cell_map entries.

Usage:
    python generate_ipr_templates.py [path/to/DPM_4.2_Annotated.xlsx]

Download: https://www.eba.europa.eu/risk-and-data-analysis/reporting-frameworks/reporting-framework-42
"""

import sys, re, os
import openpyxl
from pathlib import Path

EXCEL_PATH  = sys.argv[1] if len(sys.argv) > 1 else "DPM_4.2_Annotated.xlsx"
OUTPUT_PATH = Path("engine/definitions/ipr_templates.py")
MODULE_KEY  = "SEPA_IPR"

if not os.path.exists(EXCEL_PATH):
    print(f"ERROR: {EXCEL_PATH} not found.")
    print("Download EBA RF 4.2 technical package and place the DPM Excel here.")
    sys.exit(1)

print(f"Reading {EXCEL_PATH} …")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

concepts: dict[tuple, str] = {}
for tpl in ["S 01.01", "S 01.02", "S 02.01", "S 02.02", "S 03.00", "S 04.00"]:
    sheet_name = tpl.replace(" ", "_")
    if sheet_name not in wb.sheetnames:
        print(f"WARNING: sheet {sheet_name} not found in workbook")
        continue
    ws = wb[sheet_name]

    # Find the row containing numeric column codes (e.g. '0010', '0020') and the 'Rows' starting marker.
    col_header_row = None
    tool_row = None
    for row in ws.iter_rows(values_only=True):
        if row and isinstance(row[0], str) and row[0].strip() == "Rows":
            tool_row = row
            break
        col_header_row = row

    if not tool_row or not col_header_row:
        print(f"WARNING: no column header/Rows section in sheet {sheet_name}")
        continue

    col_positions = {}
    for i, val in enumerate(col_header_row):
        if isinstance(val, str) and val.strip().isdigit() and len(val.strip()) == 4:
            col_positions[val.strip()] = i

    row_code_index = 2

    # traverse data rows after (and including) the 'Rows' marker row
    data_started = False
    for row in ws.iter_rows(values_only=True):
        if not data_started:
            if row == tool_row:
                data_started = True
            else:
                continue
        if not row or not row[row_code_index]:
            continue

        r_code = str(row[row_code_index]).strip()
        if not r_code.isdigit() or len(r_code) != 4:
            continue

        # row is in the template rows for this table
        template_rows = [r for r, _ in __import__("engine.definitions.ipr_templates", fromlist=["TEMPLATE_STRUCTURE"]).TEMPLATE_STRUCTURE[tpl][0]]
        if r_code not in template_rows:
            continue

        for c_code, _ in __import__("engine.definitions.ipr_templates", fromlist=["TEMPLATE_STRUCTURE"]).TEMPLATE_STRUCTURE[tpl][1]:
            if c_code not in col_positions:
                continue
            v = row[col_positions[c_code]]
            if not v:
                continue
            match = re.search(r"(\d{4,})", str(v))
            if not match:
                continue
            concepts[(tpl, r_code, c_code)] = f"eba_met:mi{match.group(1)}"


print(f"Extracted {len(concepts)} SEPA_IPR concepts")

# Update ipr_templates.py by regenerating the IPR_TEMPLATES block
import engine.definitions.ipr_templates as ipr_def

# Build cell map by template from extracted concepts
cell_maps = {}
for tpl_code in set(t for t, _, _ in concepts):
    tpl_concepts = { (r, c): v for (t, r, c), v in concepts.items() if t == tpl_code }
    cell_maps[tpl_code] = tpl_concepts

# Read original template file
src_lines = OUTPUT_PATH.read_text(encoding="utf-8").splitlines()
start_idx = next(i for i, line in enumerate(src_lines) if line.strip().startswith("IPR_TEMPLATES"))
# Keep everything before IPR_TEMPLATES
new_lines = src_lines[:start_idx]
new_lines.append("IPR_TEMPLATES = {")
for tpl, (rows, cols) in ipr_def.TEMPLATE_STRUCTURE.items():
    new_lines.append(f"    \"{tpl}\": {{")
    new_lines.append(f"        \"template_code\": \"{tpl}\",")
    new_lines.append(f"        \"xbrl_filing_indicator\": \"{ipr_def.INDICATOR_MAP[tpl]}\",")
    new_lines.append(f"        \"file_name\": \"{ipr_def.TEMPLATE_FILE_NAMES[tpl]}\",")
    new_lines.append(f"        \"rows\": [")
    for r, label in rows:
        new_lines.append(f"            (\"{r}\", \"{label}\"),")
    new_lines.append("        ],")
    new_lines.append(f"        \"cols\": [")
    for c, label in cols:
        new_lines.append(f"            (\"{c}\", \"{label}\"),")
    new_lines.append("        ],")
    new_lines.append("        \"cell_map\": {")
    for (r, c), val in sorted(cell_maps.get(tpl, {}).items()):
        new_lines.append(f"            (\"{r}\", \"{c}\"): \"{val}\",")
    new_lines.append("        },")
    new_lines.append("    },")
new_lines.append("}")

OUTPUT_PATH.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
print(f"✅ ipr_templates.py updated with {len(concepts)} concept entries")
